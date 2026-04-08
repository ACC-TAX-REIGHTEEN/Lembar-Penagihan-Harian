import sys
import traceback
import openpyxl
import gspread
from google.oauth2.service_account import Credentials

config = {}
current_key = None
try:
    with open('piutang.conf', 'r') as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            if line.startswith('[') and line.endswith(']'):
                current_key = line[1:-1]
            else:
                if current_key and current_key not in config:
                    config[current_key] = line
except Exception as e:
    print(f"--> Gagal membaca piutang.conf: {e}")

try:
    wb = openpyxl.load_workbook('Print_AR_temp.xlsx', data_only=True)
    ws = wb.active
    
    data_to_insert = []
    for row in range(1, ws.max_row + 1):
        col_a = ws.cell(row=row, column=1).value
        if col_a is not None and str(col_a).strip() != "":
            baris_baru = [
                config.get('PERUSAHAAN', ''),
                str(ws.cell(row=row, column=1).value) if ws.cell(row=row, column=1).value is not None else "",
                config.get('DIVISI', ''),
                config.get('TANGGAL', ''),
                config.get('INPUT', ''),
                ws.cell(row=row, column=2).value if ws.cell(row=row, column=2).value is not None else "",
                ws.cell(row=row, column=3).value if ws.cell(row=row, column=3).value is not None else "",
                ws.cell(row=row, column=4).value if ws.cell(row=row, column=4).value is not None else "",
                ws.cell(row=row, column=5).value if ws.cell(row=row, column=5).value is not None else "",
                ws.cell(row=row, column=6).value if ws.cell(row=row, column=6).value is not None else "",
                ws.cell(row=row, column=7).value if ws.cell(row=row, column=7).value is not None else "",
                ws.cell(row=row, column=8).value if ws.cell(row=row, column=8).value is not None else "",
                ws.cell(row=row, column=9).value if ws.cell(row=row, column=9).value is not None else "",
                ws.cell(row=row, column=10).value if ws.cell(row=row, column=10).value is not None else ""
            ]
            data_to_insert.append(baris_baru)
            
    if not data_to_insert:
        print("--> Tidak ada data yang ditemukan untuk disisipkan.")
        sys.exit()
        
    scopes = [
        'https://www.googleapis.com/auth/spreadsheets',
        'https://www.googleapis.com/auth/drive'
    ]
    
    credentials = Credentials.from_service_account_file('credentials.json', scopes=scopes)
    gc = gspread.authorize(credentials)
    
    spreadsheet_id = 'YOUR SPREADSHEETS ID'
    
    try:
        sh = gc.open_by_key(spreadsheet_id)
    except gspread.exceptions.SpreadsheetNotFound:
        print("--> Error: Spreadsheets tidak ditemukan. Pastikan email Service Account sudah diberi akses Editor.")
        sys.exit()
        
    try:
        worksheet = sh.get_worksheet_by_id(YOUR ID SPREADSHEETS)
    except gspread.exceptions.WorksheetNotFound:
        print("--> Error: Worksheet dengan ID tersebut tidak ditemukan.")
        sys.exit()
    
    semua_nilai = worksheet.get_all_values()
    total_baris = len(semua_nilai)
    
    baris_sisip = total_baris - 1
    if baris_sisip < 1:
        baris_sisip = 1
        
    print(f"--> Menyiapkan {len(data_to_insert)} baris untuk disisipkan pada baris ke-{baris_sisip}...")
    
    worksheet.insert_rows(data_to_insert, row=baris_sisip, value_input_option='USER_ENTERED')
    
    print("--> Proses selesai, data berhasil disisipkan ke Spreadsheets dengan format bawaan.")
    
except Exception as e:
    print(f"--> Error detail:\n{traceback.format_exc()}")